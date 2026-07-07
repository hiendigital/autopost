"""Module 1: Data & Config Manager (local Excel via openpyxl).

Owns the `keywords.xlsx` workbook: the `Data` sheet (15-column keyword table) and the
`Cấu Hình` settings sheet. Also merges non-secret defaults from `settings.json`.
Secrets are never read from or written to Excel — see `src/env.py`.
"""
from __future__ import annotations

import json
import logging
import os
from datetime import datetime

from openpyxl import Workbook, load_workbook

logger = logging.getLogger(__name__)

# 21-column schema (1-indexed columns A..U). Columns 1-8 are INPUT (filled by the user
# via keyword research); columns 9-14 are written back by the AI; 15-21 are runtime status.
COLUMNS = [
    "Tên nhóm",                             # 1  keyword-group name (input)
    "Từ khóa chính",                        # 2  main/focus keyword (input, required)
    "Toàn bộ từ khóa cùng nhóm",            # 3  all keywords in the group (input)
    "Lượt tìm",                             # 4  search volume (input, informational)
    "Độ cạnh tranh",                        # 5  competition score (input, informational)
    "Ý định tìm kiếm của người dùng",       # 6  search intent (input, required)
    "Bài viết sẽ nhắm đến đối tượng",       # 7  target audience (input, required)
    "Chuyên mục",                           # 8  category (input, required)
    "Title Website",                        # 9  SEO title -> RankMath title (AI writes back)
    "Meta description",                     # 10 meta description -> RankMath (AI writes back)
    "Đoạn Sapo website",                    # 11 sapo/intro paragraph (AI writes back)
    "H1 của website",                       # 12 H1 = post title (AI writes back)
    "Nội dung website",                     # 13 HTML body content (AI writes back)
    "Nội dung Fanpage",                     # 14 social caption (AI writes back)
    "Trạng thái",                           # 15 status: Pending|Processing|Success|Failed
    "Link Website",                         # 16 published post URL
    "Link Fanpage",                         # 17 published FB URL
    "Link Ảnh Đại Diện",                    # 18 featured image URL
    "Link Ảnh Kèm Theo",                    # 19 in-content image URLs (comma separated)
    "Thời gian đăng",                       # 20 publish datetime
    "Nhật ký lỗi",                          # 21 error log
]
COL_INDEX = {name: i + 1 for i, name in enumerate(COLUMNS)}
STATUS_COL = COL_INDEX["Trạng thái"]

# Human-friendly keys for the row dict returned by fetch_next_pending_row.
KEY_MAP = {
    "group_name": "Tên nhóm",
    "keyword": "Từ khóa chính",
    "all_keywords": "Toàn bộ từ khóa cùng nhóm",
    "search_volume": "Lượt tìm",
    "competition": "Độ cạnh tranh",
    "intent": "Ý định tìm kiếm của người dùng",
    "audience": "Bài viết sẽ nhắm đến đối tượng",
    "category": "Chuyên mục",
    "seo_title": "Title Website",
    "meta_description": "Meta description",
    "sapo": "Đoạn Sapo website",
    "h1": "H1 của website",
    "html_content": "Nội dung website",
    "fanpage_content": "Nội dung Fanpage",
    "status": "Trạng thái",
    "link_website": "Link Website",
    "link_fanpage": "Link Fanpage",
    "featured_image_url": "Link Ảnh Đại Diện",
    "content_image_urls": "Link Ảnh Kèm Theo",
    "published_at": "Thời gian đăng",
    "error_log": "Nhật ký lỗi",
}


class ConfigManager:
    def __init__(self, settings_path: str = "settings.json"):
        with open(settings_path, "r", encoding="utf-8") as fh:
            self.settings = json.load(fh)
        self.data_file = self.settings["paths"]["data_file"]
        self.data_sheet = self.settings["sheets"]["data_sheet"]
        self.config_sheet = self.settings["sheets"]["config_sheet"]

    # ---------------------------------------------------------------- config
    def load_configurations(self) -> dict:
        """Merge settings.json defaults with overrides from the 'Cấu Hình' sheet."""
        cfg = json.loads(json.dumps(self.settings))  # deep copy
        if not os.path.exists(self.data_file):
            logger.warning("Data file %s not found; using settings.json only.", self.data_file)
            return cfg

        wb = load_workbook(self.data_file, read_only=True, data_only=True)
        if self.config_sheet not in wb.sheetnames:
            wb.close()
            return cfg
        ws = wb[self.config_sheet]
        overrides: dict[str, str] = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            key = str(row[0]).strip()
            val = row[1] if len(row) > 1 else None
            if val is not None:
                overrides[key] = val
        wb.close()

        self._apply_config_overrides(cfg, overrides)
        return cfg

    @staticmethod
    def _apply_config_overrides(cfg: dict, overrides: dict) -> None:
        prompts = cfg.setdefault("prompts", {})
        if "system_prompt_web_seo" in overrides:
            prompts["system_prompt_web_seo"] = overrides["system_prompt_web_seo"]
        if "system_prompt_social" in overrides:
            prompts["system_prompt_social"] = overrides["system_prompt_social"]
        if "image_style_prompt" in overrides:
            prompts["image_style_prompt"] = overrides["image_style_prompt"]
        if "banned_words" in overrides:
            raw = str(overrides["banned_words"])
            prompts["banned_words"] = [w.strip() for w in raw.split(",") if w.strip()]

    # ------------------------------------------------------------- data rows
    def fetch_next_pending_row(self) -> dict | None:
        """Return the first row whose status is 'Pending', or None."""
        if not os.path.exists(self.data_file):
            raise FileNotFoundError(
                f"{self.data_file} not found. Run `python main.py init-data` first."
            )
        wb = load_workbook(self.data_file, data_only=True)
        ws = wb[self.data_sheet]
        result = None
        for r in range(2, ws.max_row + 1):
            status = ws.cell(row=r, column=STATUS_COL).value
            if status is not None and str(status).strip().lower() == "pending":
                result = self._row_to_dict(ws, r)
                break
        wb.close()
        return result

    def fetch_facebook_retry_rows(self, include_skipped: bool = False) -> list[dict]:
        """Rows where WordPress is published but Facebook is not, needing a FB-only retry.

        Criteria: status == 'Success', Link Fanpage empty, and at least one stored image URL.
        By default only rows whose error log mentions Facebook are returned (the genuine
        "web posted but FB errored" case). Pass include_skipped=True to also include rows
        that were posted while Facebook was disabled (no error, just never sent to FB).
        """
        if not os.path.exists(self.data_file):
            raise FileNotFoundError(
                f"{self.data_file} not found. Run `python main.py init-data` first."
            )
        wb = load_workbook(self.data_file, data_only=True)
        ws = wb[self.data_sheet]
        rows: list[dict] = []
        for r in range(2, ws.max_row + 1):
            status = ws.cell(row=r, column=STATUS_COL).value
            if status is None or str(status).strip().lower() != "success":
                continue
            fb_link = ws.cell(row=r, column=COL_INDEX["Link Fanpage"]).value
            if fb_link is not None and str(fb_link).strip():
                continue  # already on Facebook
            featured = ws.cell(row=r, column=COL_INDEX["Link Ảnh Đại Diện"]).value
            content = ws.cell(row=r, column=COL_INDEX["Link Ảnh Kèm Theo"]).value
            if not (featured or content):
                continue  # nothing to post
            error_log = ws.cell(row=r, column=COL_INDEX["Nhật ký lỗi"]).value
            if not include_skipped:
                if not (error_log and "facebook" in str(error_log).lower()):
                    continue
            rows.append(self._row_to_dict(ws, r))
        wb.close()
        return rows

    def _row_to_dict(self, ws, r: int) -> dict:
        data = {"row_index": r}
        for key, col_name in KEY_MAP.items():
            val = ws.cell(row=r, column=COL_INDEX[col_name]).value
            data[key] = val if val is not None else ""
        return data

    def update_row_data(self, row_index: int, data_to_update: dict) -> None:
        """Write human-keyed fields back to the given row and save."""
        wb = load_workbook(self.data_file)
        ws = wb[self.data_sheet]
        for key, value in data_to_update.items():
            col_name = KEY_MAP.get(key, key)
            if col_name not in COL_INDEX:
                logger.warning("Unknown column key '%s' ignored.", key)
                continue
            ws.cell(row=row_index, column=COL_INDEX[col_name]).value = value
        wb.save(self.data_file)
        wb.close()

    def set_status(self, row_index: int, status: str) -> None:
        self.update_row_data(row_index, {"status": status})

    def log_error_to_sheet(self, row_index: int, error_message: str) -> None:
        self.update_row_data(
            row_index,
            {"status": "Failed", "error_log": str(error_message)[:2000]},
        )

    # ------------------------------------------------------------ scaffolding
    def create_sample_workbook(self) -> str:
        """Generate a fresh keywords.xlsx (Data + Cấu Hình) with sample rows."""
        os.makedirs(os.path.dirname(self.data_file) or ".", exist_ok=True)
        wb = Workbook()

        ws = wb.active
        ws.title = self.data_sheet
        ws.append(COLUMNS)
        # Columns 1-8 are INPUT (from keyword research). The 6 AI-write-back columns
        # (Title/Meta/Sapo/H1/Nội dung/Fanpage) are left blank; only status is set to Pending.
        n_ai = 6  # Title Website .. Nội dung Fanpage
        n_tail = len(COLUMNS) - STATUS_COL  # Link Website .. Nhật ký lỗi
        samples = [
            ["Chăm sóc da mùa hè", "cách chăm sóc da mùa hè",
             "cách chăm sóc da mùa hè, dưỡng da mùa hè, skincare mùa hè", 1600, 12,
             "Tìm hướng dẫn chăm sóc, bảo vệ da khỏi nắng nóng và dầu nhờn mùa hè.",
             "Nữ giới 20-35 tuổi quan tâm làm đẹp", "Làm đẹp"],
            ["Mua nhà lần đầu", "kinh nghiệm mua nhà lần đầu",
             "kinh nghiệm mua nhà lần đầu, lưu ý khi mua nhà, mua nhà trả góp", 880, 20,
             "Tìm kinh nghiệm, lưu ý và các bước cần biết khi mua nhà lần đầu.",
             "Vợ chồng trẻ chuẩn bị mua nhà", "Bất động sản"],
            ["Tiết kiệm điện tại nhà", "mẹo tiết kiệm điện tại nhà",
             "mẹo tiết kiệm điện, cách giảm tiền điện, tiết kiệm điện mùa hè", 720, 8,
             "Tìm mẹo thực tế giúp giảm hóa đơn tiền điện hàng tháng tại nhà.",
             "Gia đình muốn giảm chi phí sinh hoạt", "Đời sống"],
        ]
        for row in samples:
            ws.append(row + [""] * n_ai + ["Pending"] + [""] * n_tail)
        for col in range(1, len(COLUMNS) + 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 22

        cfg = wb.create_sheet(self.config_sheet)
        cfg.append(["Khóa cấu hình", "Giá trị"])
        p = self.settings["prompts"]
        cfg_rows = [
            ["system_prompt_web_seo", p["system_prompt_web_seo"]],
            ["system_prompt_social", p["system_prompt_social"]],
            ["image_style_prompt", p["image_style_prompt"]],
            ["banned_words", ", ".join(p["banned_words"])],
        ]
        for row in cfg_rows:
            cfg.append(row)
        cfg.column_dimensions["A"].width = 26
        cfg.column_dimensions["B"].width = 70

        wb.save(self.data_file)
        wb.close()
        logger.info("Created sample workbook at %s", self.data_file)
        return self.data_file


def now_string() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")
